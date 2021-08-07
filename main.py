import msvcrt
import os
import sys
import time
import openpyxl
from Tools.demo.beer import n
from selenium import webdriver
from selenium.webdriver.common import keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.expected_conditions import presence_of_element_located
from msvcrt import getch


if os.path.isdir('img'):
    pass
else:
    os.mkdir('img')


book = openpyxl.open("articul.xlsx")
sheet = book.active  # взять первый активный лист
# print(sheet["A1"].value)  # взять из ячейки
# print(sheet[1][0].value)  # взять из первой строки первого столбца


# <editor-fold desc="Поиск первой пустой ячейки в колонке B">
def finde_empty_cell_B():
    try:
        for row in range(1, sheet.max_row+1):
            cell_B = sheet[row][1].value
            if cell_B == None:
                break
    except:
        row = 1
    return row
# </editor-fold>

# print(finde_empty_cell_B())
# <editor-fold desc="парсинг">
# options = webdriver.ChromeOptions()
# options.add_argument('headless')
# driver = webdriver.Chrome(chrome_options=options)
driver = webdriver.Chrome()
driver.get('https://www.kromschroeder.de/marketing/adlatus/profi/profi.html?sprache=1')
time.sleep(3)
def pars():
    row_art = finde_empty_cell_B()

    elem = driver.find_element(By.CSS_SELECTOR, 'input#id_suche_eingabe')
    j = 1
    for row in range(row_art, sheet.max_row + 1):
        j = j + 1
        if j < 10:
            articl = sheet[row][0].value
            elem.send_keys(articl)
            elem.send_keys(Keys.ENTER)
            time.sleep(3)
            try:
                name = driver.find_element(By.CSS_SELECTOR, 'div#id_filter_liste_div').text  #.find_element(By.CSS_SELECTOR )
                img = driver.find_element(By.CSS_SELECTOR, 'iframe#id_zeichnung')
                img.screenshot(f'img\{name}.png')

            except:
                name = 'Ничего не найдено'

            sheet.cell(row=row, column=2).value = name
            book.save('articul.xlsx')
            elem.clear()

            print(f'{articl} {name}')
        else:
            print('Тестовый режим.')
            print('Обсудить детали напишите мне в кворке или на jon_vy@mail.ru')
            print('')

            break
# </editor-fold>
pars()

'''запись в ячейку'''
# c1 = sheet.cell(row=2, column=2)
# c1.value = "ANKIT"
# book.save(f"{os.environ['PYTHONPATH']}\\articul.xlsx")

# home = os.environ['PYTHONPATH']
# print(home)
driver.close()
driver.quit()
key = msvcrt.getch()
if key==b'\r':
    sys.exit(n)